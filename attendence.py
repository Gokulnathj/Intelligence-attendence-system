# function for face detection with mtcnn
from PIL import Image
from numpy import asarray
import numpy as npy
from os import listdir
from numpy import load
from numpy import expand_dims
from numpy import savez_compressed
from keras.models import load_model
from sklearn.preprocessing import LabelEncoder
from sklearn.preprocessing import Normalizer
from sklearn.svm import SVC
from matplotlib import pyplot 
import datetime
from openpyxl import load_workbook

# extract a single face from a given photograph
def extract_face(filename, required_size=(160, 160)):
	# load image from file
	image = Image.open(filename)
	# convert to RGB, if needed
	image = image.convert('RGB')
	# convert to array
	face = asarray(image)
	# resize pixels to the model size
	image = Image.fromarray(face)
	image = image.resize(required_size)
	face_array = asarray(image)
	return face_array
def load_faces(directory):
	faces = list()
	# enumerate files
	for filename in listdir(directory):
		# path
		path = directory + filename
		# get face
		face = extract_face(path)
		# store
		faces.append(face)
	return faces
# load a dataset that contains one subdir for each class that in turn contains images
def load_dataset(directory):
	X= list()
	# load all faces in the subdirectory
	faces = load_faces(directory)
	# store
	X.extend(faces)
	return asarray(X)

# get the face embedding for one face
def get_embedding(model, face_pixels):
	# scale pixel values
	face_pixels = face_pixels.astype('float32')
	# standardize pixel values across channels (global)
	mean, std = face_pixels.mean(), face_pixels.std()
	face_pixels = (face_pixels - mean) / std
	# transform face into one sample
	samples = expand_dims(face_pixels, axis=0)
	# make prediction to get embedding
	yhat = model.predict(samples)
	return yhat[0]


testX = load_dataset(r'C:\Users\raisw\Documents\yoloface-master\dataset/')
print(testX.shape)
# save arrays to one file in compressed format
npy.savez_compressed('test-faces-dataset.npz', testX )

# load the face dataset
data = load('test-faces-dataset.npz')
testX = data['arr_0']
print('Loaded: ', testX.shape)
# load the facenet model
model = load_model('facenet_keras.h5')
print('Loaded Model')
# convert each face in the test set to an embedding
newTestX = list()
for face_pixels in testX:
	embedding = get_embedding(model, face_pixels)
	newTestX.append(embedding)
newTestX = asarray(newTestX)
print(newTestX.shape)
# save arrays to one file in compressed format
savez_compressed('test-faces-embeddings.npz', newTestX )

# load faces
data = load('test-faces-dataset.npz')
testX_faces = data['arr_0']
# load face embeddings
data = load(r"C:\Users\raisw\Documents\training dataset\training-embeddings.npz")
trainX, trainy = data['arr_0'], data['arr_1']
data = load('test-faces-embeddings.npz')
testX = data['arr_0']
# normalize input vectors
in_encoder = Normalizer(norm='l2')
trainX = in_encoder.transform(trainX)
testX = in_encoder.transform(testX)
# label encode targets
out_encoder = LabelEncoder()
out_encoder.fit(trainy)
trainy = out_encoder.transform(trainy)
# fit model
model = SVC(kernel='linear', probability=True)
model.fit(trainX, trainy)
names=[]
# test model on a random example from the test dataset
for selection in range(testX.shape[0]):
    random_face_pixels = testX_faces[selection]
    random_face_emb = testX[selection]
    # prediction for the face
    samples = expand_dims(random_face_emb, axis=0)
    yhat_class = model.predict(samples)
    yhat_prob = model.predict_proba(samples)
    # get name
    class_index = yhat_class[0]
    class_probability = yhat_prob[0,class_index] * 100
    predict_names = out_encoder.inverse_transform(yhat_class)
    print('Predicted: %s (%.3f)' % (predict_names[0], class_probability))
    names.append(predict_names[0])
    # plot for fun
    pyplot.imshow(random_face_pixels)
    title = '%s (%.3f)' % (predict_names[0], class_probability)
    pyplot.title(title)
    pyplot.show()

#updating attendance for recognised faces   
month=int(str(datetime.datetime.now().date())[5:7])
wb = load_workbook('sample.xlsx')
ws = wb.worksheets[month]
present_date=int(str(datetime.datetime.now().date())[8:])+1
for j in range(2,13):
    if ws.cell(j,1).value in names:
        ws.cell(j,present_date).value='p'
    else:
        ws.cell(j,present_date).value='a'
# Save the file
wb.save("sample.xlsx")
print("attendance marked successfully for ",len(names), "students")








